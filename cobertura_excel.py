#%%
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

df = pd.read_csv("output/cobertura_influenza_2024_rm.csv")

nombre_grupo = {
    "ninos_6m_5basico":             "Niños 6m - 5° básico",
    "ninos_6_10_anios":             "Niños 6 a 10 años",
    "adultos_60_mas":               "Adultos 60+",
    "cronicos_11_59":               "Crónicos 11-59 años",
    "embarazadas":                  "Embarazadas",
    "estrategia_capullo":           "Estrategia Capullo",
    "salud_privado":                "P. Salud Privado",
    "salud_publico":                "P. Salud Público",
    "trab_educacion":               "Trab. Educación",
    "trab_avicolas_cerdos":         "Trab. Avícolas/Cerdos",
    "otras_prioridades":            "Otras Prioridades",
    "cuidadores_adulto_mayor_eleam":"Cuidadores/ELEAM",
}

tipo_grupo = {
    "Niños 6m - 5° básico":     "RESIDENCIA",
    "Niños 6 a 10 años":        "OCURRENCIA",
    "Adultos 60+":              "RESIDENCIA",
    "Crónicos 11-59 años":      "RESIDENCIA",
    "Embarazadas":              "RESIDENCIA",
    "Estrategia Capullo":       "RESIDENCIA",
    "P. Salud Privado":         "OCURRENCIA",
    "P. Salud Público":         "OCURRENCIA",
    "Trab. Educación":          "OCURRENCIA",
    "Trab. Avícolas/Cerdos":    "OCURRENCIA",
    "Otras Prioridades":        "OCURRENCIA",
    "Cuidadores/ELEAM":         "OCURRENCIA",
}

df["grupo"] = df["grupo"].map(nombre_grupo)
df = df.sort_values(["COMUNA", "grupo"]).reset_index(drop=True)

pct_pivot = df.pivot(index="COMUNA", columns="grupo", values="cobertura_pct")
vac_pivot = df.pivot(index="COMUNA", columns="grupo", values="vacunados")
den_pivot = df.pivot(index="COMUNA", columns="grupo", values="denominador")

# ── TOTAL POR COMUNA (fila → suma vacunados / suma denominador) ───────────────
vac_pivot["TOTAL COMUNA"]  = vac_pivot.sum(axis=1)
den_pivot["TOTAL COMUNA"]  = den_pivot.sum(axis=1)
pct_pivot["TOTAL COMUNA"]  = (vac_pivot["TOTAL COMUNA"] / den_pivot["TOTAL COMUNA"] * 100).round(2)
tipo_grupo["TOTAL COMUNA"] = ""

# ── TOTAL REGIÓN METROPOLITANA (fila al fondo) ────────────────────────────────
total_vac = df.groupby("grupo")["vacunados"].sum()
total_den = df.groupby("grupo")["denominador"].sum()
total_pct = (total_vac / total_den * 100).round(2)

pct_pivot.loc["TOTAL REGIÓN METROPOLITANA"] = total_pct
vac_pivot.loc["TOTAL REGIÓN METROPOLITANA"] = vac_pivot.loc[vac_pivot.index != "TOTAL REGIÓN METROPOLITANA"].sum()
den_pivot.loc["TOTAL REGIÓN METROPOLITANA"] = den_pivot.loc[den_pivot.index != "TOTAL REGIÓN METROPOLITANA"].sum()

# Recalcular total RM para TOTAL COMUNA
vac_pivot.loc["TOTAL REGIÓN METROPOLITANA", "TOTAL COMUNA"] = vac_pivot.loc[
    vac_pivot.index != "TOTAL REGIÓN METROPOLITANA", "TOTAL COMUNA"].sum()
den_pivot.loc["TOTAL REGIÓN METROPOLITANA", "TOTAL COMUNA"] = den_pivot.loc[
    den_pivot.index != "TOTAL REGIÓN METROPOLITANA", "TOTAL COMUNA"].sum()
pct_pivot.loc["TOTAL REGIÓN METROPOLITANA", "TOTAL COMUNA"] = round(
    vac_pivot.loc["TOTAL REGIÓN METROPOLITANA", "TOTAL COMUNA"] /
    den_pivot.loc["TOTAL REGIÓN METROPOLITANA", "TOTAL COMUNA"] * 100, 2)

# ── ESTILOS ───────────────────────────────────────────────────────────────────

AZUL_OSC  = "1F4E79"
AZUL_MED  = "2E75B6"
AZUL_CLAR = "BDD7EE"
GRIS_CLAR = "F2F2F2"
BLANCO    = "FFFFFF"
VERDE     = "C6EFCE"
NARANJ    = "FFEB9C"
VERDE_OSC = "375623"

def estilo_header(cell, bg=AZUL_OSC, color=BLANCO, size=11):
    cell.font      = Font(bold=True, color=color, name="Arial", size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def borde_fino():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def escribir_tabla(ws, pivot, titulo, fmt_num, start_row=1):
    grupos  = list(pivot.columns)
    comunas = list(pivot.index)
    ncols   = len(grupos)

    # Fila 1: título
    ws.merge_cells(start_row=start_row, start_column=1,
                   end_row=start_row, end_column=ncols + 1)
    tc = ws.cell(start_row, 1, titulo)
    estilo_header(tc, bg=AZUL_OSC, size=12)
    ws.row_dimensions[start_row].height = 22

    # Fila 2: header grupos
    estilo_header(ws.cell(start_row + 1, 1, "Comuna"), bg=AZUL_MED)
    for j, g in enumerate(grupos, 2):
        bg_h = "375623" if g == "TOTAL COMUNA" else AZUL_MED
        estilo_header(ws.cell(start_row + 1, j, g), bg=bg_h)
    ws.row_dimensions[start_row + 1].height = 40

    # Fila 3: tipo de comuna
    c_tipo = ws.cell(start_row + 2, 1, "Tipo comuna")
    c_tipo.font      = Font(bold=True, name="Arial", size=9, italic=True)
    c_tipo.fill      = PatternFill("solid", start_color=GRIS_CLAR)
    c_tipo.alignment = Alignment(horizontal="center", vertical="center")
    c_tipo.border    = borde_fino()
    for j, g in enumerate(grupos, 2):
        tipo = tipo_grupo.get(g, "")
        c    = ws.cell(start_row + 2, j, tipo)
        if g == "TOTAL COMUNA":
            bg_t = "D9E1F2"
        else:
            bg_t = VERDE if tipo == "RESIDENCIA" else NARANJ
        c.font      = Font(bold=True, name="Arial", size=9)
        c.fill      = PatternFill("solid", start_color=bg_t)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = borde_fino()
    ws.row_dimensions[start_row + 2].height = 18

    # Filas de datos
    for i, comuna in enumerate(comunas):
        row      = start_row + 3 + i
        es_total_rm = comuna == "TOTAL REGIÓN METROPOLITANA"
        bg_fila  = GRIS_CLAR if i % 2 == 0 else BLANCO

        c_com = ws.cell(row, 1, comuna)
        c_com.font      = Font(bold=True, name="Arial", size=10,
                               color="FFFFFF" if es_total_rm else "000000")
        c_com.fill      = PatternFill("solid", start_color=AZUL_OSC if es_total_rm else AZUL_CLAR)
        c_com.alignment = Alignment(vertical="center")
        c_com.border    = borde_fino()

        for j, g in enumerate(grupos, 2):
            val = pivot.loc[comuna, g]
            c   = ws.cell(row, j)
            c.value         = round(float(val), 2) if pd.notna(val) else "S/D"
            c.number_format = fmt_num

            if es_total_rm:
                c.fill = PatternFill("solid", start_color=AZUL_OSC)
                c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            elif g == "TOTAL COMUNA":
                c.fill = PatternFill("solid", start_color="D9E1F2")
                c.font = Font(name="Arial", size=10, bold=True)
            else:
                c.fill = PatternFill("solid", start_color=bg_fila)
                c.font = Font(name="Arial", size=10)

            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = borde_fino()

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions[get_column_letter(len(grupos) + 1)].width = 18  # TOTAL COMUNA más ancha
    for j in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16

    return start_row + 3 + len(comunas)

# ── WORKBOOK ──────────────────────────────────────────────────────────────────

wb = Workbook()

ws1 = wb.active
ws1.title = "Cobertura %"
ws1.freeze_panes = "B4"
last = escribir_tabla(ws1, pct_pivot, "COBERTURA VACUNACIÓN INFLUENZA 2024 — Región Metropolitana (%)", '0.00"%"')
data_range = f"B4:{get_column_letter(len(pct_pivot.columns))}{last - 1}"
ws1.conditional_formatting.add(
    data_range,
    ColorScaleRule(
        start_type="num", start_value=0,  start_color="F8696B",
        mid_type="num",   mid_value=80,   mid_color="FFEB84",
        end_type="num",   end_value=100,  end_color="63BE7B",
    )
)

ws2 = wb.create_sheet("Vacunados")
ws2.freeze_panes = "B4"
escribir_tabla(ws2, vac_pivot, "VACUNADOS INFLUENZA 2024 — Región Metropolitana", '#,##0')

ws3 = wb.create_sheet("Denominador")
ws3.freeze_panes = "B4"
escribir_tabla(ws3, den_pivot, "POBLACIÓN OBJETIVO INFLUENZA 2024 — Región Metropolitana", '#,##0')

wb.save("output/cobertura_influenza_2024_rm.xlsx")
print("✅ Guardado: output/cobertura_influenza_2024_rm.xlsx")
# %%
